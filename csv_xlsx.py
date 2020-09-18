

from openpyxl import load_workbook,Workbook
import csv

f = open("marks.csv","r")

wb = Workbook
ws = wb.active


for no,line in enumerate(csv.reader(f)):
    print(no,line)
    for rowno in len(line):
        ws.cell(row=line,column=1,value=line)
        for columnno in range(max(no)):
            ws.cell(row=1,column=columnno,value=line)

wb.save("marks_xl.xlsx")
f.close()
