from openpyxl import load_workbook

wb = load_workbook("sample1.xlsx")

print(wb.sheetnames)

ws = wb["students"]

for row in ws.values:
    print(row)


ws2 = wb["marks"]

for row in ws2.values:
    print(row)

print(ws["A1"].value)
print(ws.cell(row=2,column=1).value)
