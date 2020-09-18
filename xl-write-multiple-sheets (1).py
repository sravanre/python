from openpyxl import Workbook

wb = Workbook()

# create sheet1
ws1 = wb.create_sheet("marks")
ws1.append([89,78,56,90,67])
ws1.append([89,98,56,90,67])
ws1.append([89,78,66,90,67])


# create sheet2
ws2 = wb.create_sheet("students")
ws2.append(["hari",90,89])
ws2.append(["ravi",78,89])
ws2.append(["siri",87,89])


wb.save("sample1.csv")
